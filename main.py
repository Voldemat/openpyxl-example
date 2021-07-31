import json

from openpyxl import load_workbook
from openpyxl.styles import Border, Alignment, Side, Font

# get styles config
with open("configs/styles.json", "r", encoding = 'UTF-8') as styles_file:
	STYLES = json.load(styles_file)
	

# create cells styles...
font = Font(**STYLES['font'])
border = Border(*list(map(lambda side: Side(**side), STYLES['border'].values())))
alignment = Alignment(**STYLES['alignment'])
	
def make_table(worksheet, records) -> None:
	workzone = worksheet.iter_rows(min_row = 2, min_col = 2, max_row = 2 + len(records))

	for record, row in zip(records, workzone):
		# get row - list of cells and record - tuple values
		for index, value in enumerate(record):
			# define cell object
			cell = row[index]

			# set value
			cell.value = value

			# set all styles...
			cell.font = font
			cell.border = border
			cell.alignment = alignment


	return None

def main() -> None:
	with open("configs/records.json", "r", encoding = 'UTF-8') 	 as test_data_file,\
		 open("configs/config.script.json", "r", encoding = 'UTF-8') as config_file:
		
		RECORDS = json.load(test_data_file)
		CONFIG = json.load(config_file)
		
	workbook = load_workbook(CONFIG['inputFile'])
	
	make_table(workbook.active, RECORDS)

	workbook.save(CONFIG['outputFile'])

	return None



if __name__ == '__main__':
	main()